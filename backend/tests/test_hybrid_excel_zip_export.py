"""HYB-5 acceptance gate (part 2): hybrid Excel/ZIP export -- real
workflow run + checkpoint decision + runner-uploaded evidence must
appear in both export formats, the ZIP manifest must link every hybrid
entity by id, evidence bytes must be verified against their checksum,
and the original 7 Track A sheets/behaviors must be untouched."""
import io
import json
import zipfile

import openpyxl
from fastapi.testclient import TestClient

from app.main import app

PNG_BYTES = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x02\x00\x00\x00\x02\x08\x02\x00\x00\x00\xfd\xd4\x9as"
    b"\x00\x00\x00\x0cIDATx\x9cc\xf8\xcf\xc0\xc0\xc0\xc0\x00\x00\x00\x06\x00\x03\xfa\xd0\x7f\xe6"
    b"\x00\x00\x00\x00IEND\xaeB`\x82"
)


def _fresh_client():
    c = TestClient(app)
    c.headers.update({"Origin": "http://localhost:5173"})
    return c


def _setup_hybrid_run_with_checkpoint(auth_client, slug):
    wf = auth_client.post(f"/api/{slug}/workflows", json={"name": "export wf"}).json()
    rev = auth_client.post(f"/api/{slug}/workflows/{wf['id']}/revisions", json={"revision_label": "v1"}).json()
    cp = auth_client.post(
        f"/api/{slug}/workflows/{wf['id']}/revisions/{rev['id']}/steps",
        json={"step_type": "MANUAL_CHECKPOINT", "description": "export cp", "checkpoint_instructions": "verify"},
    ).json()
    auth_client.post(f"/api/{slug}/workflows/{wf['id']}/revisions/{rev['id']}/publish")

    suite = auth_client.post(f"/api/{slug}/suites", json={"name": "Export Suite"}).json()
    suite_rev = auth_client.post(f"/api/{slug}/suites/{suite['id']}/revisions", json={"revision_label": "rv1"}).json()
    auth_client.post(
        f"/api/{slug}/revisions/{suite_rev['id']}/cases",
        json={"checkpoint_code": "EXPORT-001", "title": "c", "action_md": "a", "expected_result_md": "e"},
    )
    auth_client.post(f"/api/{slug}/suites/{suite['id']}/revisions/{suite_rev['id']}/publish")
    cycle = auth_client.post(
        f"/api/{slug}/cycles",
        json={"suite_id": suite["id"], "script_revision_id": suite_rev["id"], "name": "export cycle", "environment": "test"},
    ).json()
    result_id = auth_client.get(f"/api/{slug}/cycles/{cycle['id']}/results").json()[0]["id"]

    queued = auth_client.post(
        f"/api/{slug}/workflow-runs", json={"workflow_revision_id": rev["id"], "cycle_test_result_id": result_id}
    ).json()
    run_id = queued["id"]
    token_resp = auth_client.post("/api/runner-tokens", json={"label": "export-runner"})
    token = token_resp.json()["token"]
    runner = _fresh_client()
    runner.headers.update({"X-Runner-Token": token})
    claim = runner.post(f"/api/{slug}/workflow-runs/claim").json()
    lease_token = claim["lease_token"]

    sr_cp = runner.post(f"/api/{slug}/workflow-runs/{run_id}/step-runs", json={"workflow_step_id": cp["id"], "lease_token": lease_token}).json()
    upload = runner.post(
        f"/api/{slug}/workflow-runs/{run_id}/evidence",
        params={"lease_token": lease_token, "step_run_id": sr_cp["id"]},
        files={"file": ("export.png", PNG_BYTES, "image/png")},
    )
    assert upload.status_code == 200, upload.text
    evidence_id = upload.json()["id"]

    runner.post(
        f"/api/{slug}/workflow-runs/{run_id}/events",
        json={"event_type": "CHECKPOINT_WAITING", "actor_type": "RUNNER", "lease_token": lease_token, "payload_json": f'{{"step_id":{cp["id"]}}}'},
    )
    decide = auth_client.post(
        f"/api/{slug}/workflow-runs/{run_id}/checkpoint-decision",
        json={"workflow_step_id": cp["id"], "status": "PASS", "actual_result_md": "ok", "evidence_ids": [evidence_id]},
    )
    assert decide.status_code == 200, decide.text
    decision_id = decide.json()["id"]

    resume = runner.post(f"/api/{slug}/workflow-runs/{run_id}/checkpoint-resume", json={"workflow_step_id": cp["id"], "lease_token": lease_token})
    assert resume.status_code == 200, resume.text
    runner.post(f"/api/{slug}/workflow-runs/{run_id}/complete", json={"status": "PASSED", "lease_token": lease_token})

    return {
        "cycle_id": cycle["id"],
        "run_id": run_id,
        "workflow_id": wf["id"],
        "revision_id": rev["id"],
        "step_id": cp["id"],
        "evidence_id": evidence_id,
        "decision_id": decision_id,
    }


def test_hybrid_excel_sheets_contain_real_run_data(auth_client, project_slug):
    slug = project_slug
    ctx = _setup_hybrid_run_with_checkpoint(auth_client, slug)

    r = auth_client.get(f"/api/{slug}/cycles/{ctx['cycle_id']}/export/excel")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))

    runs_sheet = wb["Workflow Runs"]
    run_ids_in_sheet = [row[0].value for row in runs_sheet.iter_rows(min_row=2)]
    assert ctx["run_id"] in run_ids_in_sheet

    decisions_sheet = wb["Checkpoint Decisions"]
    decision_rows = [row for row in decisions_sheet.iter_rows(min_row=2, values_only=True)]
    assert any(row[0] == ctx["decision_id"] and row[4] == "PASS" for row in decision_rows)

    timing_sheet = wb["Timing Trends"]
    timing_rows = [row for row in timing_sheet.iter_rows(min_row=2, values_only=True)]
    assert any(row[0] == ctx["run_id"] for row in timing_rows)

    steps_sheet = wb["Workflow Steps"]
    step_descriptions = [row[4] for row in steps_sheet.iter_rows(min_row=2, values_only=True)]
    assert "export cp" in step_descriptions


def test_hybrid_zip_manifest_links_every_entity_and_verifies_checksums(auth_client, project_slug):
    slug = project_slug
    ctx = _setup_hybrid_run_with_checkpoint(auth_client, slug)

    r = auth_client.get(f"/api/{slug}/cycles/{ctx['cycle_id']}/export/zip")
    assert r.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(r.content))
    manifest = json.loads(zf.read("manifest.json"))

    hybrid = manifest["hybrid"]
    assert any(w["id"] == ctx["workflow_id"] for w in hybrid["workflows"])
    assert any(rv["id"] == ctx["revision_id"] for rv in hybrid["workflow_revisions"])
    assert any(s["id"] == ctx["step_id"] for s in hybrid["workflow_steps"])
    run_entry = next(r for r in hybrid["workflow_runs"] if r["id"] == ctx["run_id"])
    assert run_entry["status"] == "PASSED"
    assert run_entry["timing"]["run_id"] == ctx["run_id"]
    decision_entry = next(d for d in hybrid["checkpoint_decisions"] if d["id"] == ctx["decision_id"])
    assert decision_entry["status"] == "PASS"
    assert decision_entry["resume_authorized"] is True

    # Evidence entry carries real hybrid provenance links and passed its
    # own checksum verification -- real bytes, not a substituted URL.
    evidence_entry = next(e for e in manifest["evidence"] if e["evidence_id"] == ctx["evidence_id"])
    assert evidence_entry["workflow_run_id"] == ctx["run_id"]
    assert evidence_entry["checkpoint_decision_id"] == ctx["decision_id"]
    assert evidence_entry["checksum_verified"] is True
    assert evidence_entry["missing"] is False
    assert evidence_entry["filename"] is not None

    # The evidence bytes are genuinely present in the archive (not a
    # presigned URL substitute) and match the original PNG exactly.
    archived_bytes = zf.read(evidence_entry["filename"])
    assert archived_bytes == PNG_BYTES

    # Sanity: the xlsx is also present in the archive (existing behavior).
    assert any(name.endswith(".xlsx") for name in zf.namelist())
