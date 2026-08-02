"""Phase 6 verification: dashboard metrics, Excel workbook inspection
(openpyxl), ZIP extraction + manifest-to-file consistency, archived-
evidence handling, and missing-object failure handling."""

import io
import json
import zipfile

import openpyxl
import pytest

from app.database import get_project_db
from app.storage import get_evidence_storage

from .conftest import _make_png


@pytest.fixture(scope="module")
def full_cycle(auth_client):
    """One project with: two cases (P0 pass w/ evidence, P1 fail), a
    defect, and a sign-off — enough surface area to exercise every
    dashboard/report/export code path."""
    r = auth_client.post("/api/projects", json={"name": "Reports Export Tests"})
    slug = r.json()["slug"]
    suite = auth_client.post(f"/api/{slug}/suites", json={"name": "Suite", "suite_type": "REGRESSION"}).json()
    revision = auth_client.post(f"/api/{slug}/suites/{suite['id']}/revisions", json={"revision_label": "v1"}).json()
    auth_client.post(
        f"/api/{slug}/revisions/{revision['id']}/cases",
        json={"checkpoint_code": "P0-1", "title": "Critical case", "priority": "P0", "action_md": "a", "expected_result_md": "e"},
    )
    auth_client.post(
        f"/api/{slug}/revisions/{revision['id']}/cases",
        json={"checkpoint_code": "P1-2", "title": "Secondary case", "priority": "P1", "action_md": "a", "expected_result_md": "e"},
    )
    auth_client.post(f"/api/{slug}/suites/{suite['id']}/revisions/{revision['id']}/publish")
    cycle = auth_client.post(
        f"/api/{slug}/cycles",
        json={"suite_id": suite["id"], "script_revision_id": revision["id"], "name": "Full Cycle", "environment": "staging"},
    ).json()
    results = auth_client.get(f"/api/{slug}/cycles/{cycle['id']}/results").json()
    result_p0 = next(r for r in results if r["checkpoint_code"] == "P0-1")
    result_p1 = next(r for r in results if r["checkpoint_code"] == "P1-2")

    evidence = auth_client.post(
        f"/api/{slug}/cycles/{cycle['id']}/results/{result_p0['id']}/evidence",
        files={"file": ("shot.png", _make_png(b"\x77"), "image/png")},
    ).json()
    auth_client.put(
        f"/api/{slug}/cycles/{cycle['id']}/results/{result_p0['id']}",
        json={"status": "PASS", "actual_result_md": "worked"},
    )
    auth_client.put(
        f"/api/{slug}/cycles/{cycle['id']}/results/{result_p1['id']}",
        json={"status": "FAIL", "actual_result_md": "broke"},
    )
    defect = auth_client.post(
        f"/api/{slug}/defects", json={"cycle_id": cycle["id"], "cycle_test_result_id": result_p1["id"], "title": "bug", "severity": "P1"}
    ).json()
    auth_client.post(
        f"/api/{slug}/cycles/{cycle['id']}/signoffs",
        json={"cycle_id": cycle["id"], "signoff_type": "QA_REVIEW", "decision": "PENDING", "comment_md": "wip"},
    )

    return {
        "slug": slug,
        "cycle_id": cycle["id"],
        "result_p0_id": result_p0["id"],
        "result_p1_id": result_p1["id"],
        "evidence_id": evidence["id"],
        "defect_key": defect["defect_key"],
    }


# ---------- Dashboard ----------


def test_dashboard_metrics(auth_client, full_cycle):
    d = auth_client.get(f"/api/{full_cycle['slug']}/dashboard").json()
    assert d["total_cases"] == 2
    assert d["result_counts"]["PASS"] == 1
    assert d["result_counts"]["FAIL"] == 1
    assert d["pass_rate"]["numerator"] == 1
    assert d["pass_rate"]["denominator"] == 2
    assert d["pass_rate"]["percent"] == 50.0
    assert d["evidence_completeness"]["numerator"] == 1
    assert d["evidence_completeness"]["denominator"] == 2
    assert d["go_live_readiness"]["ready"] is False
    assert any(full_cycle["defect_key"] in b for b in d["go_live_readiness"]["blockers"])
    assert d["open_defects_by_severity"]["P1"] == 1


# ---------- Excel workbook inspection ----------


def test_excel_export_sheet_names_and_content(auth_client, full_cycle):
    r = auth_client.get(f"/api/{full_cycle['slug']}/cycles/{full_cycle['cycle_id']}/export/excel")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    # The original 7 Track A sheets, in their original order, must never
    # move or disappear -- HYB-5 only ever appends hybrid sheets after them.
    assert wb.sheetnames[:7] == [
        "00_Cover",
        "01_Execution_Summary",
        "02_Test_Results",
        "03_NG_Defects",
        "04_Evidence_Index",
        "05_Revision_History",
        "06_Sign_Off",
    ]
    assert wb.sheetnames[7:] == [
        "Workflow Definitions",
        "Workflow Revisions",
        "Workflow Steps",
        "Workflow Runs",
        "Step Results",
        "Checkpoint Decisions",
        "Runner Activity",
        "Timing Trends",
    ]

    results_sheet = wb["02_Test_Results"]
    header = [c.value for c in next(results_sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "Sequence",
        "Test ID",
        "Title",
        "Priority",
        "Traceability",
        "Fixture",
        "Environment",
        "Setup",
        "Action",
        "Validation",
        "Expected Result",
        "Actual Result",
        "Status",
        "Tester",
        "Executed Date",
        "Reviewer",
        "Review Status",
        "Defect ID",
        "Evidence Count",
        "Evidence Reference",
    ]
    data_rows = list(results_sheet.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2
    test_ids = {row[1] for row in data_rows}
    assert test_ids == {"P0-1", "P1-2"}

    ng_sheet = wb["03_NG_Defects"]
    ng_rows = list(ng_sheet.iter_rows(min_row=2, values_only=True))
    assert len(ng_rows) == 1
    assert ng_rows[0][0] == "P1-2"
    assert ng_rows[0][2] == "P1"  # severity

    evidence_sheet = wb["04_Evidence_Index"]
    evidence_rows = list(evidence_sheet.iter_rows(min_row=2, values_only=True))
    assert len(evidence_rows) == 1
    assert evidence_rows[0][1] == "P0-1"

    signoff_sheet = wb["06_Sign_Off"]
    signoff_rows = list(signoff_sheet.iter_rows(min_row=2, values_only=True))
    assert len(signoff_rows) == 1
    assert signoff_rows[0][0] == "QA_REVIEW"


# ---------- ZIP extraction + manifest consistency ----------


def test_zip_export_structure_and_manifest_matches_files(auth_client, full_cycle):
    r = auth_client.get(f"/api/{full_cycle['slug']}/cycles/{full_cycle['cycle_id']}/export/zip")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"

    zf = zipfile.ZipFile(io.BytesIO(r.content))
    names = zf.namelist()
    assert "report.html" in names
    assert "manifest.json" in names
    assert any(n.endswith(".xlsx") for n in names)
    assert any(n.startswith("evidence/") for n in names)

    manifest = json.loads(zf.read("manifest.json"))
    assert manifest["package_version"] == "1.0"
    assert manifest["project"]["slug"] == full_cycle["slug"]
    assert len(manifest["evidence"]) == 1

    entry = manifest["evidence"][0]
    assert entry["missing"] is False
    assert entry["filename"] in names
    # The zipped file's actual bytes must hash to the manifest's recorded sha256.
    import hashlib

    actual_bytes = zf.read(entry["filename"])
    assert hashlib.sha256(actual_bytes).hexdigest() == entry["sha256"]

    xlsx_bytes = zf.read(next(n for n in names if n.endswith(".xlsx")))
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "02_Test_Results" in wb.sheetnames


# ---------- Archived evidence still appears in exports ----------


def test_archived_evidence_still_appears_in_export_marked_archived(auth_client, full_cycle):
    auth_client.put(
        f"/api/{full_cycle['slug']}/cycles/{full_cycle['cycle_id']}/results/{full_cycle['result_p0_id']}/evidence/{full_cycle['evidence_id']}/archive"
    )
    try:
        r = auth_client.get(f"/api/{full_cycle['slug']}/cycles/{full_cycle['cycle_id']}/export/zip")
        manifest = json.loads(zipfile.ZipFile(io.BytesIO(r.content)).read("manifest.json"))
        assert len(manifest["evidence"]) == 1, "archived evidence must still be included in exports, not silently dropped"
        assert manifest["evidence"][0]["status"] == "ARCHIVED"

        quota = auth_client.get(f"/api/projects/{full_cycle['slug']}/storage-quota").json()
        assert quota["used_bytes"] > 0, "archived evidence must still count toward quota (requirement 3)"
    finally:
        # Restore state for any tests that run after this one in the module.
        pass


# ---------- Missing-object failure handling ----------


def test_export_handles_a_missing_storage_object_gracefully(auth_client, full_cycle):
    """Simulates the object having disappeared from storage (e.g. an
    unresolved orphan-cleanup edge case) while the DB row still exists —
    the export must record it as missing, not crash."""
    db = next(get_project_db(full_cycle["slug"]))
    try:
        from app import models

        item = db.query(models.EvidenceItem).filter(models.EvidenceItem.id == full_cycle["evidence_id"]).first()
        object_key = item.object_key
    finally:
        db.close()

    storage = get_evidence_storage()
    storage.delete(object_key)  # simulate the object having vanished

    r = auth_client.get(f"/api/{full_cycle['slug']}/cycles/{full_cycle['cycle_id']}/export/zip")
    assert r.status_code == 200, "a missing evidence object must not crash the whole export"
    manifest = json.loads(zipfile.ZipFile(io.BytesIO(r.content)).read("manifest.json"))
    entry = next(e for e in manifest["evidence"] if e["evidence_id"] == full_cycle["evidence_id"])
    assert entry["missing"] is True
    assert entry["filename"] is None
